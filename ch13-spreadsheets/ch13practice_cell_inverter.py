#! python3

# Spreadsheet Cell Inverter
# Write a program to invert the row and column of the cells in the 
# spreadsheet. For example, the value at row 5, column 3 will be at row 3, 
# column 5 (and vice versa). This should be done for all cells in the 
# spreadsheet.

import logging
import openpyxl
import sys
import os

from openpyxl.utils import coordinate_to_tuple
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format="%(levelname)s: %(message)s")
logging.disable(logging.CRITICAL)

if len(sys.argv) != 2:
    print(f"Usage: {os.path.basename(__file__)} <spreadsheet to modify>")
    print(f"E.g., {os.path.basename(__file__)} my_spreadsheet.xlsx")
    sys.exit()

target_file = sys.argv[1]
wb = openpyxl.load_workbook(target_file)
sheet = wb.active

# Get max dimensions.
max_row = sheet.max_row 
max_col = sheet.max_column
extent = sheet.cell(row=max_row, column=max_col)

# Copies existing data into tuple "snapshot" and clears cell.
data_range = tuple(sheet[f'A1:{extent.coordinate}'])
snapshot = {}

for data_row in data_range:
    for data_cell in data_row:
        snapshot[data_cell.coordinate] = data_cell.value
        logging.info(f"data_cell: {data_cell}, value: {data_cell.value}")
        sheet[f'{data_cell.coordinate}'] = None     # Clears cell.

logging.info(f"snapshot of table data: {snapshot}")

# Transposes data.
for key in snapshot:
    new_col, new_row = coordinate_to_tuple(key)
    sheet.cell(row=new_row, column=new_col, value=snapshot[key])

    logging.info(f"moved cell {key} to {get_column_letter(new_col)}{new_row}, value: {snapshot[key]}")

# Saves workbook.
wb.save(f'transposed_{target_file}')
