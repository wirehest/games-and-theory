#! python3

# Write a program that performs the tasks of the previous program in reverse 
# order: the program should open a spreadsheet and write the cells of column A 
# into one text file, the cells of column B into another text file, and so on.

import logging
import openpyxl

from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format="%(levelname)s: %(message)s")
#logging.disable(logging.CRITICAL)

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active

# Gathers data in spreadsheet by columns.
data_columns = sheet.columns

# Counter for filenaming.
col_counter = 1

for column in data_columns:
    filename = f'column-{get_column_letter(col_counter)}.txt'
    open_txt = open(filename, 'w')
    logging.info(f"writing to {filename}")

    for row in column:
        logging.info(f"\twriting row.value: {row.value}")
        open_txt.write(str(row.value) + "\n")

    col_counter += 1

open_txt.close()
