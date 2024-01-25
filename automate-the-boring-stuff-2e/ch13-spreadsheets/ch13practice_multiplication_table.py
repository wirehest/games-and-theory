#! python3

# Multiplication Table Maker
# Create a program multiplicationTable.py that takes a number N from the
# command line and creates an N×N multiplication table in an Excel 
# spreadsheet. 

import logging
import openpyxl
import sys
import os

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')
#logging.disable(logging.CRITICAL)

# Enforce variable input from commandline.
if len(sys.argv) < 2:
    print(f"Usage: {os.path.basename(__file__)} <positive integer>")
    print(f"E.g., {os.path.basename(__file__)} 9")
    sys.exit()

factor = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active

# Create row/column labels.
for i in range(1, factor + 1):
    xlabel = sheet.cell(row=(i + 1), column=1, value=i)
    xlabel.font = Font(bold=True)
    ylabel = sheet.cell(row=1, column=(i + 1), value=i)
    ylabel.font = Font(bold=True)
    logging.info(f"create labels for: {xlabel.coordinate}, {ylabel.coordinate}")

# Formula generator.
end_cell = sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate 
table_cells = tuple(sheet['B2':end_cell])
logging.info(f"\ntable_cells: {table_cells}")

for rows in table_cells:
    for cell in rows:
        logging.info(f"cell: {cell}")
        x_position = get_column_letter(cell.column)
        y_position = cell.row
        sheet[cell.coordinate] = f'={x_position}1*A{y_position}'
        logging.info(f"{cell}, {cell.value}")

# Save workbook. 
wb.save('multiplication_table.xlsx')

