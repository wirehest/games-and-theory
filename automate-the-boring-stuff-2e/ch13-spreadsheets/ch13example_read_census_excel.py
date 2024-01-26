#! python3
# read_census_excel
# Tabulates population and number of census tracts.

import openpyxl
import pprint

wb = openpyxl.load_workbook('censuspopdata.xlsx')

sheet = wb['Population by Census Tract']
county_data = {}

# TODO fill in county data with each county's

print("Reading rows...")

for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state   = sheet['B' + str(row)].value
    county  = sheet['C' + str(row)].value
    pop     = sheet['D' + str(row)].value

    # Make sure the key for this state exists.
    county_data.setdefault(state, {})
    
    # Make sure the key for this county in this state exists.
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census tract, so increment by one.
    county_data[state][county]['tracts'] += 1

    # Increase county pop by the pop in this census tract.
    county_data[state][county]['pop'] += int(pop)

for row in range(2, sheet.max_row + 1):
    pass

# TODO: Open a new text file and write the contents of county_data to it.
print("Writing results...")
result_file = open('census2010.py', 'w')
result_file.write(f"all_data = {pprint.pformat(county_data)}")
result_file.close()
print("Done.")